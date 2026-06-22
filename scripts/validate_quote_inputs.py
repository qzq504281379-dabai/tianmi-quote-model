#!/usr/bin/env python3
"""Pre-check Tianmi quote CSV inputs against product and price data."""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook


ALIASES = {
    "product_id": ["product_id", "产品编号", "产品编码", "商品编号", "编码"],
    "product_name": ["product_name", "产品名称", "商品名称", "名称"],
    "model": ["model", "型号", "规格型号"],
    "version": ["version", "版本", "款式"],
    "specification": ["specification", "规格", "参数"],
    "unit_price": ["unit_price", "单价", "采购单价", "价格", "报价"],
    "requested_item": ["requested_item", "客户需求", "产品", "产品名称", "需求产品"],
    "quantity": ["quantity", "数量", "qty"],
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [{k.strip(): (v or "").strip() for k, v in row.items()} for row in csv.DictReader(handle)]


def read_xlsx(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    rows: list[dict[str, str]] = []
    for sheet in workbook.worksheets:
        raw_rows = sheet.iter_rows(values_only=True)
        headers = next(raw_rows, None)
        if not headers:
            continue
        clean_headers = [str(header).strip() if header is not None else "" for header in headers]
        for raw_row in raw_rows:
            record = {"_sheet": sheet.title}
            has_value = False
            for index, header in enumerate(clean_headers):
                if not header:
                    continue
                cell_value = raw_row[index] if index < len(raw_row) else None
                text = "" if cell_value is None else str(cell_value).strip()
                if text:
                    has_value = True
                if header in record and text:
                    record[f"{header}_{index + 1}"] = text
                else:
                    record[header] = text
            if has_value:
                rows.append(record)
    return rows


def read_table(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx(path)
    raise ValueError(f"Unsupported file type: {path}")


def value(row: dict[str, str], key: str) -> str:
    for alias in ALIASES[key]:
        if alias in row and row[alias]:
            return row[alias]
    return ""


def norm(text: str) -> str:
    return "".join(text.lower().split())


def item_key(row: dict[str, str]) -> tuple[str, str, str]:
    return (norm(value(row, "product_id")), norm(value(row, "product_name")), norm(value(row, "model")))


def item_label(row: dict[str, str]) -> str:
    parts = [value(row, "product_id"), value(row, "product_name"), value(row, "model"), value(row, "version")]
    label = " ".join(part for part in parts if part)
    sheet = row.get("_sheet", "")
    return f"{sheet}: {label}" if sheet else label


def candidate_matches(request: dict[str, str], catalog: list[dict[str, str]]) -> list[dict[str, str]]:
    requested = norm(value(request, "requested_item"))
    requested_model = norm(value(request, "model"))
    requested_version = norm(value(request, "version"))
    matches = []
    for product in catalog:
        pid, name, model = item_key(product)
        version = norm(value(product, "version"))
        if requested_version and requested_version != version:
            continue
        if requested_model and requested_model == model:
            matches.append(product)
        elif requested and requested in {pid, name, model}:
            matches.append(product)
    if len(matches) > 1 and not requested_version:
        capacitor_matches = [match for match in matches if norm(value(match, "version")) == norm("电容版")]
        if len(capacitor_matches) == 1:
            return capacitor_matches
    return matches


def has_price(product: dict[str, str], prices: list[dict[str, str]]) -> bool:
    product_keys = item_key(product)
    for price in prices:
        price_keys = item_key(price)
        price_value = value(price, "unit_price")
        if price_value and any(a and a == b for a, b in zip(product_keys, price_keys)):
            return True
    return False


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate quote request items against catalog and price CSV files.")
    parser.add_argument("--catalog", required=True, type=Path)
    parser.add_argument("--prices", required=True, type=Path)
    parser.add_argument("--request", required=True, type=Path)
    args = parser.parse_args()

    catalog = read_table(args.catalog)
    prices = read_table(args.prices)
    requests = read_table(args.request)
    problems = 0

    for index, request in enumerate(requests, start=1):
        requested = value(request, "requested_item") or value(request, "model") or f"row {index}"
        quantity = value(request, "quantity")
        matches = candidate_matches(request, catalog)

        if not quantity:
            print(f"[MISSING_QUANTITY] row {index}: {requested}")
            problems += 1
        if not matches:
            print(f"[NO_CATALOG_MATCH] row {index}: {requested}")
            problems += 1
            continue
        if len(matches) > 1:
            options = "; ".join(filter(None, [item_label(m) for m in matches]))
            print(f"[AMBIGUOUS_PRODUCT] row {index}: {requested} -> {options}")
            problems += 1
            continue
        if not has_price(matches[0], prices):
            print(f"[NO_PRICE] row {index}: {requested}")
            problems += 1

    if problems:
        print(f"Validation finished with {problems} issue(s). Ask for clarification or updated source files before quoting.")
        return 1

    print("Validation passed: every requested row has one catalog match, a price, and a quantity.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
